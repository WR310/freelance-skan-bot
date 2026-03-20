import asyncio
import logging
import xml.etree.ElementTree as ET
import re
import aiohttp
import sqlite3
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from google import genai
from aiogram import Bot, Dispatcher, types, F
from aiogram.filters import Command, CommandObject
from aiogram.client.default import DefaultBotProperties
from aiogram.client.session.aiohttp import AiohttpSession
from aiogram.enums import ParseMode
from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton, CallbackQuery, FSInputFile, BotCommand
from playwright.async_api import async_playwright
import sys

# ==========================================
# НАСТРОЙКИ 
# ==========================================
TELEGRAM_BOT_TOKEN = "8732409277:AAGEYg8ptrWGygY-EmB23rcm93gFLtWE5AU"
TELEGRAM_USER_ID = 1652878568
GEMINI_API_KEY = "AIzaSyBGkI3Pcbv8xSl8PRnYKXKHoQ19ubhvyIM"

# --- НАСТРОЙКА ПРОКСИ ---
USE_PROXY = True
# Если HTTP прокси не работает, попробуй поменять "http://" на "socks5://"
PROXY_URL = "socks5://Xm8w9UTx:eJ9AisX1@176.53.133.112:64863"

DEFAULT_KEYWORDS = ["python", "telegram", "телеграм", "парсер", "api", "скрипт", "чат-бот", "бот", "openai", "chatgpt"]

RSS_FEEDS = {
    "FL": "https://www.fl.ru/rss/all.xml"
}

CHECK_INTERVAL = 300  

# ==========================================
# ГЛОБАЛЬНЫЕ ПЕРЕМЕННЫЕ
# ==========================================
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

bot = None
client = None
dp = Dispatcher()
force_scan_event = asyncio.Event()
start_time = datetime.now() 

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36'
}

# ==========================================
# БАЗА ДАННЫХ (SQLITE)
# ==========================================
def init_db():
    conn = sqlite3.connect('scanner.db')
    cursor = conn.cursor()
    cursor.execute('''CREATE TABLE IF NOT EXISTS seen_jobs (id TEXT PRIMARY KEY)''')
    cursor.execute('''CREATE TABLE IF NOT EXISTS keywords (word TEXT PRIMARY KEY)''')
    
    cursor.execute('SELECT COUNT(*) FROM keywords')
    if cursor.fetchone()[0] == 0:
        for k in DEFAULT_KEYWORDS:
            cursor.execute('INSERT OR IGNORE INTO keywords (word) VALUES (?)', (k.lower(),))
            
    conn.commit()
    conn.close()

def is_job_seen(job_id: str) -> bool:
    conn = sqlite3.connect('scanner.db')
    cursor = conn.cursor()
    cursor.execute('SELECT 1 FROM seen_jobs WHERE id = ?', (job_id,))
    result = cursor.fetchone()
    conn.close()
    return bool(result)

def mark_job_seen(job_id: str):
    conn = sqlite3.connect('scanner.db')
    cursor = conn.cursor()
    cursor.execute('INSERT OR IGNORE INTO seen_jobs (id) VALUES (?)', (job_id,))
    conn.commit()
    conn.close()

def get_total_seen_jobs() -> int:
    conn = sqlite3.connect('scanner.db')
    cursor = conn.cursor()
    cursor.execute('SELECT COUNT(*) FROM seen_jobs')
    count = cursor.fetchone()[0]
    conn.close()
    return count

def get_keywords() -> list:
    conn = sqlite3.connect('scanner.db')
    cursor = conn.cursor()
    cursor.execute('SELECT word FROM keywords')
    words = [row[0] for row in cursor.fetchall()]
    conn.close()
    return words

def add_keyword(word: str) -> bool:
    conn = sqlite3.connect('scanner.db')
    cursor = conn.cursor()
    try:
        cursor.execute('INSERT INTO keywords (word) VALUES (?)', (word.lower(),))
        conn.commit()
        success = True
    except sqlite3.IntegrityError:
        success = False
    conn.close()
    return success

def remove_keyword(word: str) -> bool:
    conn = sqlite3.connect('scanner.db')
    cursor = conn.cursor()
    cursor.execute('DELETE FROM keywords WHERE word = ?', (word.lower(),))
    deleted = cursor.rowcount > 0
    conn.commit()
    conn.close()
    return deleted

# ==========================================
# EXCEL CRM СИСТЕМА
# ==========================================
def save_to_excel(title: str, link: str, cover_letter: str):
    file_name = "clients.xlsx"
    try:
        if not os.path.exists(file_name):
            wb = Workbook()
            ws = wb.active
            ws.title = "Лиды с бирж"
            ws.append(["Дата", "Название заказа", "Ссылка", "Сгенерированный отклик", "Статус (в работе/отказ/игнор)"])
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 40
            ws.column_dimensions['C'].width = 30
            ws.column_dimensions['D'].width = 60
            ws.column_dimensions['E'].width = 20
        else:
            wb = load_workbook(file_name)
            ws = wb.active

        date_str = datetime.now().strftime("%Y-%m-%d %H:%M")
        ws.append([date_str, title, link, cover_letter, ""])
        wb.save(file_name)
        logging.info(f"Excel: Заказ '{title[:20]}...' успешно добавлен в CRM.")
    except Exception as e:
        logging.error(f"Ошибка сохранения в Excel: {e}")

# ==========================================
# ТЕЛЕГРАМ КОМАНДЫ
# ==========================================
@dp.message(Command("start"))
async def cmd_start(message: types.Message):
    if message.from_user.id == TELEGRAM_USER_ID:
        kb = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="🔎 Искать прямо сейчас", callback_data="force_scan")],
            [InlineKeyboardButton(text="📖 Открыть справку", callback_data="show_help")]
        ])
        await message.answer(
            "🤖 <b>СКАЙНЕТ АКТИВИРОВАН</b>\n\n"
            "Добро пожаловать в панель управления автономной системой лидогенерации.\n"
            "Используйте кнопку <b>«Меню»</b> слева внизу для быстрой навигации.", 
            reply_markup=kb
        )

@dp.message(Command("help"))
async def cmd_help(message: types.Message):
    if message.from_user.id == TELEGRAM_USER_ID:
        await send_help_message(message)

async def send_help_message(message_or_callback):
    text = (
        "📖 <b>БАЗА ЗНАНИЙ СИСТЕМЫ</b>\n\n"
        "<b>Доступные команды:</b>\n"
        "🔹 /status — Вызов дашборда со статистикой\n"
        "🔹 /keys — Просмотр активных ключевых слов\n"
        "🔹 /get_crm — Скачать таблицу (Excel) с лидами\n\n"
        "<b>Управление фильтрами:</b>\n"
        "➕ <code>/add_key слово</code> — добавить фильтр\n"
        "➖ <code>/del_key слово</code> — удалить фильтр"
    )
    if isinstance(message_or_callback, types.Message):
        await message_or_callback.answer(text)
    else:
        await message_or_callback.message.answer(text)

@dp.callback_query(F.data == "show_help")
async def process_show_help(callback: CallbackQuery):
    if callback.from_user.id == TELEGRAM_USER_ID:
        await callback.answer()
        await send_help_message(callback)

@dp.message(Command("status"))
async def cmd_status(message: types.Message):
    if message.from_user.id == TELEGRAM_USER_ID:
        total_jobs = get_total_seen_jobs()
        total_keys = len(get_keywords())
        
        uptime = datetime.now() - start_time
        hours, remainder = divmod(uptime.seconds, 3600)
        minutes, _ = divmod(remainder, 60)
        uptime_str = f"{uptime.days} дн. {hours} ч. {minutes} мин."
        
        kb = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="📥 Выгрузить CRM (Excel)", callback_data="get_crm_btn")]
        ])
        
        await message.answer(
            "📊 <b>ДАШБОРД УПРАВЛЕНИЯ</b>\n"
            "━━━━━━━━━━━━━━━━━━━━\n"
            "🟢 <b>Статус:</b> Online\n"
            f"⏱ <b>Время работы:</b> {uptime_str}\n"
            f"📡 <b>Сенсоры:</b> FL, Kwork, Freelancium, Work24\n"
            f"🧠 <b>AI-Модуль:</b> Активен\n\n"
            "<b>СТАТИСТИКА БАЗЫ:</b>\n"
            f"🎯 <b>Активных фильтров:</b> {total_keys}\n"
            f"🗂 <b>Проанализировано заказов:</b> {total_jobs}\n"
            "━━━━━━━━━━━━━━━━━━━━\n"
            f"<i>Цикл сканирования: каждые {CHECK_INTERVAL // 60} мин.</i>",
            reply_markup=kb
        )

@dp.message(Command("keys"))
async def cmd_keys(message: types.Message):
    if message.from_user.id == TELEGRAM_USER_ID:
        keys = get_keywords()
        keys_text = "\n".join([f"• <code>{k}</code>" for k in keys])
        await message.answer(
            f"🔑 <b>АКТИВНЫЕ ФИЛЬТРЫ ({len(keys)}):</b>\n\n{keys_text}"
        )

@dp.message(Command("add_key"))
async def cmd_add_key(message: types.Message, command: CommandObject):
    if message.from_user.id == TELEGRAM_USER_ID:
        if not command.args:
            await message.answer("⚠️ Укажите слово.")
            return
        word = command.args.strip().lower()
        if add_keyword(word):
            await message.answer(f"✅ Фильтр <b>{word}</b> добавлен.")
        else:
            await message.answer(f"⚠️ Фильтр <b>{word}</b> уже существует.")

@dp.message(Command("del_key"))
async def cmd_del_key(message: types.Message, command: CommandObject):
    if message.from_user.id == TELEGRAM_USER_ID:
        if not command.args:
            await message.answer("⚠️ Укажите слово.")
            return
        word = command.args.strip().lower()
        if remove_keyword(word):
            await message.answer(f"🗑 Фильтр <b>{word}</b> удален.")
        else:
            await message.answer(f"⚠️ Фильтр <b>{word}</b> не найден.")

@dp.message(Command("get_crm"))
async def cmd_get_crm(message: types.Message):
    await send_crm_file(message)

@dp.callback_query(F.data == "get_crm_btn")
async def process_get_crm_btn(callback: CallbackQuery):
    if callback.from_user.id == TELEGRAM_USER_ID:
        await callback.answer("Готовлю файл...", show_alert=False)
        await send_crm_file(callback)

async def send_crm_file(message_or_callback):
    user_id = message_or_callback.from_user.id
    if user_id == TELEGRAM_USER_ID:
        file_path = "clients.xlsx"
        if os.path.exists(file_path):
            try:
                document = FSInputFile(file_path)
                msg = "📊 Ваша актуальная база лидов с фриланса."
                if isinstance(message_or_callback, types.Message):
                    await message_or_callback.answer_document(document, caption=msg)
                else:
                    await message_or_callback.message.answer_document(document, caption=msg)
            except Exception as e:
                logging.error(f"Ошибка при отправке CRM: {e}")
        else:
            err_msg = "⚠️ База CRM пока пуста."
            if isinstance(message_or_callback, types.Message):
                await message_or_callback.answer(err_msg)
            else:
                await message_or_callback.message.answer(err_msg)

@dp.callback_query(F.data == "force_scan")
async def process_force_scan(callback: CallbackQuery):
    if callback.from_user.id == TELEGRAM_USER_ID:
        await callback.answer("Запускаю сканирование...", show_alert=False)
        force_scan_event.set() 

# ==========================================
# ЛОГИКА ИИ И ПАРСИНГА
# ==========================================
async def generate_cover_letter(title: str, description: str) -> str:
    global client
    prompt = f"""
    Ты — профессиональный Python-разработчик на фрилансе. Заказ: {title}. Описание: {description}.
    ШАГ 1: Если это мусор или скам — напиши РОВНО ОДНО СЛОВО: SKIP.
    ШАГ 2: Иначе напиши короткий, уверенный отклик (4-5 предложений). Без воды.
    """
    try:
        loop = asyncio.get_running_loop()
        response = await loop.run_in_executor(None, lambda: client.models.generate_content(
            model='gemini-2.5-flash',
            contents=prompt
        ))
        await asyncio.sleep(4) 
        return response.text.strip()
    except Exception as e:
        logging.error(f"Ошибка API Gemini: {e}")
        return "⚠️ Ошибка генерации."

async def fetch_rss_feed(session: aiohttp.ClientSession, source_name: str, url: str) -> list:
    jobs = []
    try:
        async with session.get(url, headers=HEADERS, timeout=15) as response:
            if response.status != 200:
                return jobs
            content = await response.text()
            try:
                root = ET.fromstring(content)
            except ET.ParseError:
                return jobs
            for item in root.findall('.//item'):
                title = item.find('title').text if item.find('title') is not None else ""
                link = item.find('link').text if item.find('link') is not None else ""
                description = item.find('description').text if item.find('description') is not None else ""
                description = re.sub(r'<[^>]+>', '', description)
                jobs.append({"id": link, "title": f"[{source_name}] {title.strip()}", "link": link.strip(), "description": description.strip()})
    except Exception as e:
        logging.error(f"Ошибка RSS {source_name}: {e}")
    return jobs

async def fetch_kwork_jobs(browser) -> list:
    jobs = []
    try:
        page = await browser.new_page()
        logging.info("Playwright: Открываю Kwork...")
        await page.goto("https://kwork.ru/projects?c=11", timeout=60000)
        await page.wait_for_selector('.want-card', timeout=15000)
        cards = await page.query_selector_all('.want-card')
        for card in cards[:15]:
            title_el = await card.query_selector('.wants-card__header-title a')
            desc_el = await card.query_selector('.wants-card__description-text')
            if title_el and desc_el:
                title = await title_el.inner_text()
                link = await title_el.get_attribute('href')
                description = await desc_el.inner_text()
                full_link = link if link.startswith('http') else f"https://kwork.ru{link}"
                jobs.append({"id": full_link, "title": f"[Kwork] {title.strip()}", "link": full_link, "description": description.strip()})
        await page.close()
        logging.info(f"Playwright: Kwork успешно спарсен ({len(jobs)} заказов).")
    except Exception as e:
        logging.error(f"Ошибка Playwright Kwork: {e}")
    return jobs

async def fetch_freelancium_jobs(browser) -> list:
    jobs = []
    try:
        page = await browser.new_page()
        logging.info("Playwright: Открываю Freelancium...")
        await page.goto("https://freelancium.ru/projects", timeout=60000)
        await page.wait_for_selector('h2 a[href*="/project/"]', timeout=15000) 
        cards = await page.query_selector_all('a.shadow-sm.border')
        for card in cards[:15]:
            title_el = await card.query_selector('h2 a')
            desc_el = await card.query_selector('div.break-words')
            if title_el and desc_el:
                title = await title_el.inner_text()
                link = await title_el.get_attribute('href')
                description = await desc_el.inner_text()
                full_link = link if link.startswith('http') else f"https://freelancium.ru{link}"
                jobs.append({"id": full_link, "title": f"[Freelancium] {title.strip()}", "link": full_link, "description": description.strip()})
        await page.close()
        logging.info(f"Playwright: Freelancium успешно спарсен ({len(jobs)} заказов).")
    except Exception as e:
        logging.error(f"Ошибка Playwright Freelancium: {e}")
        try: await page.close() 
        except: pass
    return jobs

async def fetch_work24_jobs(browser) -> list:
    jobs = []
    try:
        page = await browser.new_page()
        logging.info("Playwright: Открываю Work24...")
        await page.goto("https://work24.ru/orders", timeout=60000) 
        await page.wait_for_selector('a.order-item__subhead__left__title__link', timeout=15000) 
        elements = await page.query_selector_all('a.order-item__subhead__left__title__link')
        for el in elements[:15]:
            title = await el.inner_text()
            link = await el.get_attribute('href')
            if title and link:
                full_link = link if link.startswith('http') else f"https://work24.ru{link}"
                jobs.append({"id": full_link, "title": f"[Work24] {title.strip()}", "link": full_link, "description": "Детали внутри карточки на сайте Work24" })
        await page.close()
        logging.info(f"Playwright: Work24 успешно спарсен ({len(jobs)} заказов).")
    except Exception as e:
        logging.error(f"Ошибка Playwright Work24: {e}")
        try: await page.close() 
        except: pass
    return jobs

async def scan_freelance_boards():
    global bot
    try:
        await bot.send_message(TELEGRAM_USER_ID, "🚀 <b>Скайнет v15.0 запущен!</b>\nПоиск инициализирован.")
    except Exception as e:
        logging.error(f"Ошибка TG: {e}")
        
    async with aiohttp.ClientSession() as session:
        while True:
            try:
                logging.info("Сканирую ленты (FL + Playwright 3x)...")
                force_scan_event.clear() 
                
                current_keys = get_keywords()
                compiled_keywords = [re.compile(rf'\b{re.escape(k)}\b', re.IGNORECASE) for k in current_keys]
                
                def contains_keywords(title: str) -> bool:
                    return any(pattern.search(title) for pattern in compiled_keywords)
                
                all_jobs = []
                
                fl_jobs = await fetch_rss_feed(session, "FL", RSS_FEEDS["FL"])
                all_jobs.extend(fl_jobs)
                
                async with async_playwright() as p:
                    browser = await p.chromium.launch(headless=True, args=['--disable-blink-features=AutomationControlled'])
                    kwork_jobs = await fetch_kwork_jobs(browser)
                    all_jobs.extend(kwork_jobs)
                    freelancium_jobs = await fetch_freelancium_jobs(browser)
                    all_jobs.extend(freelancium_jobs)
                    work24_jobs = await fetch_work24_jobs(browser)
                    all_jobs.extend(work24_jobs)
                    await browser.close()
                
                for job in all_jobs:
                    if is_job_seen(job['id']):
                        continue
                    
                    mark_job_seen(job['id'])
                    
                    if contains_keywords(job['title']):
                        cover_letter = await generate_cover_letter(job['title'], job['description'])
                        if "SKIP" in cover_letter.upper():
                            continue
                        
                        save_to_excel(job['title'], job['link'], cover_letter)
                        kb = InlineKeyboardMarkup(inline_keyboard=[[InlineKeyboardButton(text="🔗 Открыть заказ", url=job['link'])]])
                        msg = f"🔥 <b>НОВЫЙ ЛИД</b> #новый_заказ\n━━━━━━━━━━━━━━━━━━━━\n<b>Проект:</b> {job['title']}\n\n🤖 <b>AI Отклик:</b>\n<blockquote>{cover_letter}</blockquote>"
                        
                        try:
                            await bot.send_message(TELEGRAM_USER_ID, msg, reply_markup=kb)
                        except Exception as e:
                            logging.error(f"Ошибка отправки: {e}")
                
                logging.info(f"Проверка завершена. Жду {CHECK_INTERVAL} сек.")
                await asyncio.wait_for(force_scan_event.wait(), timeout=CHECK_INTERVAL)
                logging.info("⚡ Запущен принудительный поиск по кнопке!")
                
            except asyncio.TimeoutError:
                pass 
            except Exception as e:
                logging.error(f"Глобальная ошибка в цикле парсинга: {e}. Перезапуск через 10 секунд...")
                await asyncio.sleep(10)

# ==========================================
# ОСНОВНОЙ ЦИКЛ (С ПРАВИЛЬНЫМ РЕСТАРТОМ)
# ==========================================
async def main():
    global bot, client
    
    if USE_PROXY and PROXY_URL:
        logging.info("Инициализация: ПРОКСИ ВКЛЮЧЕН.")
        os.environ['http_proxy'] = PROXY_URL
        os.environ['https_proxy'] = PROXY_URL
        client = genai.Client(api_key=GEMINI_API_KEY)
        bot_session = AiohttpSession(proxy=PROXY_URL)
        bot = Bot(token=TELEGRAM_BOT_TOKEN, session=bot_session, default=DefaultBotProperties(parse_mode=ParseMode.HTML))
    else:
        logging.info("Инициализация: ПРЯМОЕ ПОДКЛЮЧЕНИЕ (БЕЗ ПРОКСИ).")
        if 'http_proxy' in os.environ: del os.environ['http_proxy']
        if 'https_proxy' in os.environ: del os.environ['https_proxy']
        client = genai.Client(api_key=GEMINI_API_KEY)
        bot = Bot(token=TELEGRAM_BOT_TOKEN, default=DefaultBotProperties(parse_mode=ParseMode.HTML))

    init_db()
    
    commands = [
        BotCommand(command="start", description="Перезапустить бота"),
        BotCommand(command="status", description="Панель управления (Дашборд)"),
        BotCommand(command="keys", description="Настройка фильтров"),
        BotCommand(command="get_crm", description="Выгрузить Excel базу"),
        BotCommand(command="help", description="База знаний (Справка)")
    ]
    await bot.set_my_commands(commands)
    
    scanner_task = asyncio.create_task(scan_freelance_boards())
    polling_task = asyncio.create_task(dp.start_polling(bot))
    
    try:
        await asyncio.gather(scanner_task, polling_task)
    finally:
        await bot.session.close()

if __name__ == "__main__":
    while True:
        try:
            asyncio.run(main())
        except KeyboardInterrupt:
            logging.info("Остановлено вручную.")
            break
        except Exception as e:
            logging.error(f"Падение Event Loop: {e}. Бронебойный рестарт через 5 сек...")
            import time
            time.sleep(5)